"""
Erstellt content-log.xlsx für Socialeap Instagram Content-Dokumentation
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import os

OUT = "/root/Social Media Posts/outputs/instagram/content-log.xlsx"

# ── Brand Colors (hex ohne #) ──────────────────────────────────────────────────
BG_DARK   = "0A0A0E"
BLUE      = "52C8F5"
ORANGE    = "F5A028"
MAGENTA   = "E63264"
OFF_WHITE = "F5F5F7"
DARK_ROW1 = "14141A"
DARK_ROW2 = "1A1A22"
TEXT_DIM  = "6B6B80"

# ── Helper fills & fonts ───────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=11, color="F5F5F7"):
    return Font(name="Calibri", bold=True, size=size, color=color)

def reg_font(size=10, color="F5F5F7"):
    return Font(name="Calibri", size=size, color=color)

def dim_font(size=10, color=TEXT_DIM):
    return Font(name="Calibri", size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_wrap():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)

def thin_border():
    side = Side(style="thin", color="2A2A35")
    return Border(left=side, right=side, top=side, bottom=side)

# ── Columns definition ─────────────────────────────────────────────────────────
COLUMNS = [
    ("ID",           8),
    ("Datum",        12),
    ("Uhrzeit",      10),
    ("Format",       14),
    ("Säule",        16),
    ("Thema",        28),
    ("Hook",         40),
    ("Caption",      55),
    ("Hashtags",     45),
    ("Slides",        8),
    ("Dateipfad",    40),
    ("Status",       14),
    ("Likes",         9),
    ("Kommentare",    9),
    ("Saves",         9),
    ("Notizen",      35),
]

# ── Data rows ──────────────────────────────────────────────────────────────────
ROWS = [
    [
        "001",
        "14.04.2026",
        "11:30",
        "Carousel",
        "Expertise",
        "5 SEO-Fehler die deine Website unsichtbar machen",
        "5 Fehler, die deine Website bei Google unsichtbar machen.",
        (
            "5 Fehler, die deine Website bei Google unsichtbar machen.\n\n"
            "Die meisten Unternehmen investieren Tausende in ihre Website.\n"
            "Und trotzdem findet sie keiner.\n\n"
            "Nicht wegen des Designs.\nNicht wegen des Textes.\n"
            "Sondern wegen 5 technischer Fehler, die kaum jemand kennt.\n\n"
            "Schreib uns, wenn einer davon bei dir zutrifft.\n"
            "Erstgespräch kostenlos — Link in Bio."
        ),
        "#socialeap #automatisierung #webentwicklung #digitalwachsen "
        "#prozessautomatisierung #seo #googleranking #websiteerstellen "
        "#onlinemarketing #leadgenerierung #webdesign #kmu #unternehmer "
        "#geschäftsführer #digitalagentur",
        7,
        "posts/2026-04-14-seo-fehler/",
        "Erstellt",
        "",
        "",
        "",
        "",
    ],
    [
        "002",
        "16.04.2026",
        "11:30",
        "Single Image",
        "Problem",
        "Manuelle Prozesse rauben Zeit",
        "Dein Team macht heute noch Dinge, die eine Maschine erledigen könnte.",
        (
            "Dein Team macht heute noch Dinge, die eine Maschine erledigen könnte.\n\n"
            "Angebote per Hand schreiben.\n"
            "Daten aus E-Mails in Excel kopieren.\n"
            "Rechnungen manuell versenden.\n\n"
            "Das kostet dich jeden Monat Stunden — manchmal Tage.\n\n"
            "Automatisierung ist keine Zukunftsmusik.\n"
            "Die Frage ist nur: Wie lange noch?\n\n"
            "Schreib uns. Wir zeigen dir, wo bei dir Zeit verloren geht."
        ),
        "#socialeap #automatisierung #prozessautomatisierung #digitalwachsen "
        "#workflowautomation #businessautomation #zeitsparen #effizienz "
        "#digitalisierung #kmu #mittelstand #unternehmer #geschäftsführer "
        "#agenturtipp #b2b",
        1,
        "posts/2026-04-16-manuelle-prozesse/",
        "Geplant",
        "",
        "",
        "",
        "",
    ],
    [
        "003",
        "17.04.2026",
        "11:00",
        "Carousel",
        "Social Proof",
        "Case Study: 170 Leads in 6 Monaten",
        "170 Leads in 6 Monaten — so haben wir das gebaut.",
        (
            "170 Leads in 6 Monaten — so haben wir das gebaut.\n\n"
            "Ein Dienstleister aus dem Mittelstand.\n"
            "Vorher: 0 Anfragen über die Website.\n"
            "Nachher: 170 qualifizierte Anfragen in 6 Monaten.\n\n"
            "Was hat sich verändert?\n"
            "→ Swipe für die komplette Aufschlüsselung."
        ),
        "#socialeap #webentwicklung #leadgenerierung #seo #digitalwachsen "
        "#onlinemarketing #googleranking #webdesign #casestudy "
        "#erfolgsgeschichte #b2b #unternehmer #geschäftsführer "
        "#digitalagentur #results",
        6,
        "posts/2026-04-17-case-study/",
        "Geplant",
        "",
        "",
        "",
        "",
    ],
    [
        "004",
        "18.04.2026",
        "11:00",
        "Single Image",
        "Behind the Scenes",
        "Tool-Tipp: n8n spart 5h/Woche",
        "Dieses Tool spart uns 5 Stunden pro Woche — und es ist kostenlos.",
        (
            "Dieses Tool spart uns 5 Stunden pro Woche — und es ist kostenlos.\n\n"
            "Wir reden von n8n.\n"
            "Einem Open-Source-Tool für Automatisierungen.\n\n"
            "Formulardaten → automatisch ins CRM.\n"
            "Neue Anfrage → direkt in Slack benachrichtigt.\n"
            "Rechnungseingang → sofort kategorisiert.\n\n"
            "Kein Code nötig. Kein Abo-Modell.\n"
            "Nur smarte Abläufe.\n\n"
            "Was läuft bei dir noch manuell?"
        ),
        "#socialeap #automatisierung #n8n #workflowautomation #zeitsparen "
        "#procesautomation #tooloftheweek #digitalisierung #kmu #startup "
        "#unternehmer #agenturtipp #b2b #nocode #effizienz",
        1,
        "posts/2026-04-18-tool-tipp/",
        "Geplant",
        "",
        "",
        "",
        "",
    ],
    [
        "005",
        "19.04.2026",
        "10:00",
        "Reel-Script",
        "CTA",
        "Erstgespräch-Angebot",
        "Wenn deine Website keine Anfragen bringt — woran liegt das wirklich?",
        (
            "Wenn deine Website keine Anfragen bringt — woran liegt das wirklich?\n\n"
            "Wir schauen es uns gemeinsam an.\n"
            "Kostenlos. 30 Minuten. Ohne Vertriebsgespräch.\n\n"
            "Link in Bio."
        ),
        "#socialeap #webentwicklung #digitalagentur #digitalwachsen "
        "#websiteerstellen #seo #leadgenerierung #onlinemarketing "
        "#unternehmer #geschäftsführer #erstgespräch #b2b "
        "#agenturtipp #automatisierung #kmu",
        0,
        "posts/2026-04-19-cta/",
        "Geplant",
        "",
        "",
        "",
        "",
    ],
]

# ── Status color mapping ───────────────────────────────────────────────────────
STATUS_COLORS = {
    "Erstellt": ("1A3A1A", "52C85F"),   # dark green bg, green text
    "Geplant":  ("1A1A3A", "52C8F5"),   # dark blue bg, blue text
    "Gepostet": ("2A1A0A", "F5A028"),   # dark orange bg, orange text
}

def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Content Log"

    # ── Sheet background (row by row since openpyxl doesn't support sheet BG) ─
    ws.sheet_properties.tabColor = BLUE

    # ── Title row ───────────────────────────────────────────────────────────────
    ws.merge_cells("A1:P1")
    title_cell = ws["A1"]
    title_cell.value = "SOCIALEAP — Instagram Content Log  |  KW 16 · 2026"
    title_cell.font  = Font(name="Calibri", bold=True, size=16, color=OFF_WHITE)
    title_cell.fill  = fill(BG_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Accent stripe
    ws.merge_cells("A2:P2")
    stripe = ws["A2"]
    stripe.fill = fill(BLUE)
    ws.row_dimensions[2].height = 4

    # ── Header row ──────────────────────────────────────────────────────────────
    header_row = 3
    ws.row_dimensions[header_row].height = 32
    for col_idx, (col_name, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font      = bold_font(11, BG_DARK)
        cell.fill      = fill(BLUE)
        cell.alignment = center()
        cell.border    = thin_border()

    # ── Data rows ────────────────────────────────────────────────────────────────
    for row_idx, row_data in enumerate(ROWS, start=4):
        bg = DARK_ROW1 if (row_idx % 2 == 0) else DARK_ROW2
        ws.row_dimensions[row_idx].height = 80

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = fill(bg)
            cell.border = thin_border()

            col_name = COLUMNS[col_idx - 1][0]

            # ID
            if col_name == "ID":
                cell.font      = bold_font(11, BLUE)
                cell.alignment = center()
            # Status
            elif col_name == "Status" and value in STATUS_COLORS:
                bg_s, fg_s = STATUS_COLORS[value]
                cell.fill      = fill(bg_s)
                cell.font      = bold_font(10, fg_s)
                cell.alignment = center()
            # Slides
            elif col_name == "Slides":
                cell.font      = bold_font(11, ORANGE)
                cell.alignment = center()
            # Format
            elif col_name == "Format":
                cell.font      = bold_font(10, ORANGE)
                cell.alignment = center()
            # Säule
            elif col_name == "Säule":
                color_map = {
                    "Expertise":        BLUE,
                    "Problem":          MAGENTA,
                    "Social Proof":     "52F5A0",
                    "Behind the Scenes":"F5E152",
                    "CTA":              ORANGE,
                }
                c = color_map.get(str(value), OFF_WHITE)
                cell.font      = bold_font(10, c)
                cell.alignment = center()
            # Caption, Hashtags, Hook, Notizen — left-aligned wrap
            elif col_name in ("Caption", "Hashtags", "Hook", "Notizen", "Thema", "Dateipfad"):
                cell.font      = reg_font(9, OFF_WHITE)
                cell.alignment = left_wrap()
            # Empty perf columns
            elif col_name in ("Likes", "Kommentare", "Saves"):
                cell.font      = dim_font(10)
                cell.alignment = center()
            # Date, time
            elif col_name in ("Datum", "Uhrzeit"):
                cell.font      = reg_font(10, OFF_WHITE)
                cell.alignment = center()
            else:
                cell.font      = reg_font(10, OFF_WHITE)
                cell.alignment = left_wrap()

    # ── Column widths ────────────────────────────────────────────────────────────
    for col_idx, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze header ────────────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Auto-filter ──────────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUMNS))}3"

    # ── Legend sheet ─────────────────────────────────────────────────────────────
    lg = wb.create_sheet("Legende")
    lg.sheet_properties.tabColor = ORANGE
    lg["A1"].value = "Socialeap Instagram — Legende & Farb-Schema"
    lg["A1"].font  = Font(bold=True, size=14, color=OFF_WHITE)
    lg["A1"].fill  = fill(BG_DARK)

    legends = [
        ("Säule", "Farbe",    "Beschreibung"),
        ("Expertise",         BLUE,    "Wissen teilen, Autorität aufbauen"),
        ("Problem",           MAGENTA, "Pain der Zielgruppe ansprechen"),
        ("Social Proof",      "52F5A0","Ergebnisse & Kundenstimmen"),
        ("Behind the Scenes", "F5E152","Team, Tools, Einblicke"),
        ("CTA",               ORANGE,  "Direkte Handlungsaufforderung"),
    ]
    for r_i, (lbl, clr, desc) in enumerate(legends, start=3):
        lg.cell(r_i, 1, lbl).font  = Font(bold=(r_i==3), size=11, color=OFF_WHITE if r_i>3 else BG_DARK)
        lg.cell(r_i, 1).fill       = fill(BG_DARK if r_i==3 else clr) if r_i>3 else fill(BLUE)
        lg.cell(r_i, 1).alignment  = center()
        lg.cell(r_i, 2).fill       = fill(clr) if r_i>3 else fill(BLUE)
        lg.cell(r_i, 3, desc).font = Font(size=10, color=OFF_WHITE)
        lg.cell(r_i, 3).fill       = fill(BG_DARK)

    for col in [1, 2, 3]:
        lg.column_dimensions[get_column_letter(col)].width = 25 if col != 2 else 12

    wb.save(OUT)
    print(f"✓  content-log.xlsx gespeichert:\n   {OUT}")

if __name__ == "__main__":
    main()
